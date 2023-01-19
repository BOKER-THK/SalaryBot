import os
import essentials
import utils
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import matplotlib
from matplotlib import pyplot as plt

matplotlib.use("agg")

excel_path = essentials.DATA_PATH + "Salaries.xlsx"
pic_path = essentials.DATA_PATH + "SalaryChart.png"


def save_excel(df):
    df.to_excel(excel_path, index_label=True)


def get_dataframe():
    if not os.path.exists(excel_path):
        df = utils.create_salary_df(None)
    else:
        df = pd.read_excel(excel_path, index_col="date")
    return df


def get_row(index, df=get_dataframe()):
    res = ""
    row = df.loc[index]
    for i in utils.columns:
        res += i + ": " + str(row[i]) + ", "
    return res[:-2]


def process_df():
    df = get_dataframe()
    df["expenses"] = df["total"] - (df["taxes"] + df["earnings"])
    df.drop("total", axis=1, inplace=True)
    df.sort_index(ascending=True, inplace=True)
    return df


def update_data(fresh_df):
    original_df = get_dataframe()
    duplicates_df = original_df.join(fresh_df, how="inner", lsuffix="_x")

    add_df = pd.concat([original_df, fresh_df.drop(index=duplicates_df.index)])
    save_excel(add_df)

    workbook = load_workbook(excel_path)
    sheet = workbook.active
    for i in range(2, sheet.max_row + 1):
        sheet[f"A{i}"].number_format = numbers.FORMAT_DATE_XLSX17
    workbook.save(excel_path)

    left_df_cols = list(map(lambda x: x + "_x", utils.columns))
    duplicates_df = duplicates_df[duplicates_df[utils.columns] != duplicates_df[left_df_cols].rename(
        dict(zip(left_df_cols, utils.columns)), axis=1)].dropna(how="all")
    duplicates_df = fresh_df.loc[duplicates_df.index]

    if len(duplicates_df.index) > 0:
        raise utils.DuplicateException("Duplicate values found in data, handler failed", duplicates_df[utils.columns])


def update_row(index, fresh_row):
    df = get_dataframe()
    df.loc[index] = fresh_row
    save_excel(df)


def produce_graph():
    df = process_df()
    colors = ['r', 'b', 'g']
    plt.figure(figsize=(8, 5))

    for (i, col) in enumerate(df.columns):
        plt.plot([], [], label=col, color=colors[i])

    plt.stackplot(df.index, df["taxes"], df["earnings"], df["expenses"])

    plt.legend()
    plt.savefig(pic_path)


def produce_report():
    df = process_df()
    palette = plt.get_cmap("Set1")

    fig, axes = plt.subplots(3, figsize=(8, 15))

    for (i, col) in enumerate(df.columns):
        axes[i].plot(df.index, df[col], color=palette(i), label=col)
        axes[i].plot(df.index, [df[col].mean()] * df.index.size, label="avg", color='k')

        axes[i].legend()
        axes[i].set_title(col)

    plt.suptitle("Report findings", fontname="fantasy", fontsize=28)
    plt.tight_layout(pad=2.5)

    plt.savefig(pic_path)


















